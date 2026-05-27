from flask import Blueprint, request, send_file
import os
import tempfile
import openpyxl
import random

excel_process_bp = Blueprint('excel_process', __name__)


def adjust_prices(ws, target_total, limit_price=False, price_min=50, price_max=95):
    rows = []
    total_rows = 0
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            rows.append(list(row))
            total_rows += 1
    
    if not rows:
        return
    
    min_ratio = price_min / 100
    max_ratio = price_max / 100
    
    for row in rows:
        original_price = row[4] if row[4] else 0
        if original_price > 0:
            ratio = random.uniform(min_ratio, max_ratio)
            new_price = original_price * ratio
            
            if original_price >= 100:
                new_price = int(round(new_price))
            else:
                new_price = round(new_price, 2)
            
            row[5] = new_price
        else:
            row[5] = 0
    
    def calculate_total():
        total = 0
        for row in rows:
            new_price = row[5] if row[5] else 0
            quantity = row[6] if row[6] else 0
            total += new_price * quantity
        return total
    
    current_total = calculate_total()
    error_threshold = target_total * 0.01
    
    while abs(current_total - target_total) > error_threshold:
        item_totals = []
        for i, row in enumerate(rows):
            new_price = row[5] if row[5] else 0
            quantity = row[6] if row[6] else 0
            item_total = new_price * quantity
            item_totals.append((item_total, i))
        
        item_totals.sort(reverse=True, key=lambda x: x[0])
        adjustment_needed = target_total - current_total
        
        for item_total, index in item_totals:
            if abs(adjustment_needed) < error_threshold:
                break
            
            row = rows[index]
            original_price = row[4] if row[4] else 0
            current_price = row[5] if row[5] else 0
            quantity = row[6] if row[6] else 0
            
            if quantity == 0:
                continue
            
            item_adjustment = adjustment_needed / quantity
            new_price = current_price + item_adjustment
            
            min_price = original_price * min_ratio
            max_price = original_price * max_ratio
            
            if original_price >= 100:
                new_price = int(round(new_price))
                min_price = int(round(min_price))
                max_price = int(round(max_price))
            else:
                new_price = round(new_price, 2)
                min_price = round(min_price, 2)
                max_price = round(max_price, 2)
            
            new_price = max(min_price, min(max_price, new_price))
            
            old_price = row[5]
            row[5] = new_price
            adjustment_needed -= (new_price - old_price) * quantity
        
        current_total = calculate_total()
    
    for i, row in enumerate(rows, start=2):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=value)
    
    final_total = calculate_total()
    print(f"目标总价: {target_total}, 调整后总价: {final_total}, 误差: {abs(final_total - target_total):.2f}")

@excel_process_bp.route('/process-excel', methods=['POST'])
def process_excel():
    try:
        if 'file' not in request.files:
            return "没有选择文件", 400
        file = request.files['file']

        if file.filename == '':
            return "没有选择文件", 400

        if not file.filename.lower().endswith('.xlsx'):
            return "不支持的文件类型。请上传 .xlsx 文件。", 400

        target_total = request.form.get('targetTotalPrice', type=float)
        if target_total is None:
            return "请输入目标总价", 400
        
        limit_price = request.form.get('limitPrice', 'false').lower() == 'true'
        price_min = request.form.get('priceMin', type=int, default=50)
        price_max = request.form.get('priceMax', type=int, default=95)

        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        file.save(temp_file.name)
        temp_file_path = temp_file.name
        temp_file.close()

        try:
            wb = openpyxl.load_workbook(temp_file_path)
            ws = wb.active

            adjust_prices(ws, target_total, limit_price, price_min, price_max)

            output_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            output_path = output_file.name
            output_file.close()
            wb.save(output_path)

            original_name = os.path.basename(file.filename)
            new_filename = f"NEW_{original_name}"

            return send_file(output_path, as_attachment=True, download_name=new_filename)

        finally:
            if os.path.exists(temp_file_path):
                os.unlink(temp_file_path)

    except Exception as e:
        return f"处理文件时出错: {str(e)}", 500