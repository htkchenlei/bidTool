"""
项目管理 API — Blueprint
招标资料关键信息提取与废标风险提示系统
"""
import os
import json
import time
import uuid
import re
import shutil
from datetime import datetime
from flask import Blueprint, jsonify, request, send_file

projects_bp = Blueprint("projects", __name__)

BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
DATA_DIR = os.path.join(BASE_DIR, "data")
UPLOADS_DIR = os.path.join(DATA_DIR, "uploads")
PARSED_DIR = os.path.join(DATA_DIR, "parsed")
EXPORTS_DIR = os.path.join(DATA_DIR, "exports")

# 确保目录存在
for d in [UPLOADS_DIR, PARSED_DIR, EXPORTS_DIR]:
    os.makedirs(d, exist_ok=True)

# 数据文件路径
PROJECTS_FILE = os.path.join(DATA_DIR, "projects.json")
PROJECT_FILES_FILE = os.path.join(DATA_DIR, "project_files.json")
EXTRACTED_FIELDS_FILE = os.path.join(DATA_DIR, "extracted_fields.json")
RISK_ITEMS_FILE = os.path.join(DATA_DIR, "risk_items.json")

# ── Markitdown 网页转 MD ─────────────────────────────────────
def convert_url_to_markdown(url: str, project_id: str = None) -> dict:
    """
    使用 markitdown 将公告网页 URL 转换为 Markdown 格式
    返回: {
        "success": bool,
        "markdown_content": str,
        "content_length": int,
        "error": str or None
    }
    """
    result = {
        "success": False,
        "markdown_content": "",
        "content_length": 0,
        "error": None
    }

    try:
        from markitdown import MarkItDown

        md_converter = MarkItDown()
        conversion_result = md_converter.convert(url)

        # 提取 Markdown 文本内容
        md_content = conversion_result.text_content if hasattr(conversion_result, 'text_content') else str(conversion_result)

        if md_content and len(md_content.strip()) > 10:
            result["success"] = True
            result["markdown_content"] = md_content.strip()
            result["content_length"] = len(md_content.strip())

            # 保存 MD 文件到 parsed 目录（供后续查看和复用）
            if project_id:
                parsed_dir = os.path.join(PARSED_DIR, project_id)
                os.makedirs(parsed_dir, exist_ok=True)
                md_file_path = os.path.join(parsed_dir, "announcement.md")
                with open(md_file_path, "w", encoding="utf-8") as f:
                    f.write(f"# 公告网页内容\n\n**来源URL**: {url}\n**转换时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n---\n\n{md_content}")
                result["md_file_path"] = md_file_path
        else:
            result["error"] = "markitdown 转换结果为空，可能网页无法正常访问或内容过少"

    except ImportError:
        result["error"] = "markitdown 库未安装，请运行: pip install markitdown"
    except Exception as e:
        result["error"] = f"markitdown 转换失败: {str(e)}"

    return result


def get_saved_markdown(project_id: str) -> str:
    """获取项目已保存的公告 MD 内容"""
    md_file_path = os.path.join(PARSED_DIR, project_id, "announcement.md")
    if os.path.exists(md_file_path):
        with open(md_file_path, "r", encoding="utf-8") as f:
            return f.read()
    return ""


# ── 字段定义 ────────────────────────────────────────────────
FIELD_DEFINITIONS = {
    # 基础信息
    "project_name": {"label": "项目名称", "group": "基础信息"},
    "project_no": {"label": "项目编号", "group": "基础信息"},
    "section_no": {"label": "标段或包号", "group": "基础信息"},
    "bidding_method": {"label": "招标方式", "group": "基础信息"},
    "procurement_type": {"label": "采购类型", "group": "基础信息"},
    "project_location": {"label": "项目地点", "group": "基础信息"},
    "announcement_url": {"label": "公告地址", "group": "基础信息"},
    # 机构与联系人
    "purchaser": {"label": "招标人或采购人", "group": "机构与联系人"},
    "agency": {"label": "招标代理机构", "group": "机构与联系人"},
    "purchaser_contact": {"label": "招标人联系人", "group": "机构与联系人"},
    "purchaser_phone": {"label": "招标人联系电话", "group": "机构与联系人"},
    "agency_contact": {"label": "代理机构联系人", "group": "机构与联系人"},
    "agency_phone": {"label": "代理机构联系电话", "group": "机构与联系人"},
    "email": {"label": "电子邮箱", "group": "机构与联系人"},
    # 金额信息
    "budget_amount": {"label": "预算金额", "group": "金额信息"},
    "max_price": {"label": "最高限价", "group": "金额信息"},
    "control_price": {"label": "招标控制价", "group": "金额信息"},
    "bid_bond_amount": {"label": "投标保证金金额", "group": "金额信息"},
    "performance_bond": {"label": "履约保证金", "group": "金额信息"},
    # 时间信息
    "announce_date": {"label": "公告发布时间", "group": "时间信息"},
    "register_start": {"label": "报名开始时间", "group": "时间信息"},
    "register_end": {"label": "报名截止时间", "group": "时间信息"},
    "clarify_end": {"label": "答疑或澄清截止时间", "group": "时间信息"},
    "bond_deadline": {"label": "保证金缴纳截止时间", "group": "时间信息"},
    "bid_deadline": {"label": "投标截止时间", "group": "时间信息"},
    "opening_time": {"label": "开标时间", "group": "时间信息"},
    # 项目建设内容
    "industry": {"label": "所属行业", "group": "项目建设内容"},
}

# 风险类别
RISK_CATEGORIES = [
    "明确废标/否决投标",
    "资格审查不通过",
    "符合性审查不通过",
    "特殊标记实质性条款",
    "不允许负偏离",
    "签字盖章要求",
    "密封和递交要求",
    "投标保证金要求",
    "报价和最高限价要求",
    "服务期限或交付期要求",
    "运维响应要求",
    "源代码、知识产权、数据归属要求",
    "等保、密评、信创、国产化要求",
    "人员配置或项目负责人要求",
    "业绩证明或案例要求",
    "其他需关注条款",
]

# 风险关键词
RISK_KEYWORDS = [
    "废标", "否决投标", "无效投标", "投标无效", "不予受理",
    "实质性响应", "实质性条款", "不得负偏离", "必须满足",
    "资格审查", "符合性审查", "星号条款", "实质性要求",
]

# 特殊符号
SPECIAL_MARKS = ["★", "*", "▲", "#", "※"]


# ── 数据加载与保存 ────────────────────────────────────────
def _load_json(filepath, default=None):
    if default is None:
        default = []
    if os.path.exists(filepath):
        try:
            with open(filepath, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return default


def _save_json(filepath, data):
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def load_projects():
    return _load_json(PROJECTS_FILE, [])


def save_projects(data):
    _save_json(PROJECTS_FILE, data)


def load_project_files():
    return _load_json(PROJECT_FILES_FILE, [])


def save_project_files(data):
    _save_json(PROJECT_FILES_FILE, data)


def load_extracted_fields():
    return _load_json(EXTRACTED_FIELDS_FILE, [])


def save_extracted_fields(data):
    _save_json(EXTRACTED_FIELDS_FILE, data)


def load_risk_items():
    return _load_json(RISK_ITEMS_FILE, [])


def save_risk_items(data):
    _save_json(RISK_ITEMS_FILE, data)


# ── 项目管理 API ────────────────────────────────────────────
@projects_bp.route("/api/projects", methods=["GET"])
def list_projects():
    """获取项目列表"""
    projects = load_projects()

    # 筛选参数
    status = request.args.get("status", "")
    keyword = request.args.get("keyword", "").lower()

    filtered = projects
    if status:
        filtered = [p for p in filtered if p.get("status") == status]
    if keyword:
        filtered = [p for p in filtered if keyword in p.get("name", "").lower()
                    or keyword in p.get("project_no", "").lower()
                    or keyword in p.get("purchaser", "").lower()]

    # 排序：按更新时间倒序
    filtered.sort(key=lambda x: x.get("updated_at", ""), reverse=True)

    return jsonify({"projects": filtered, "total": len(filtered)})


@projects_bp.route("/api/stats", methods=["GET"])
def get_stats():
    """获取工作台统计数据"""
    projects = load_projects()
    project_files = load_project_files()

    # 递归统计文件管理模块的文件总数（data/files 目录）
    FILES_DIR = os.path.join(DATA_DIR, "files")
    total_files = 0
    if os.path.exists(FILES_DIR):
        for root, dirs, files in os.walk(FILES_DIR):
            total_files += len([f for f in files if os.path.isfile(os.path.join(root, f))])

    # 统计已分析项目数（有字段提取记录的项目）
    extracted_fields = load_extracted_fields()
    analyzed_projects = len(set(
        f.get("project_id") for f in extracted_fields if f.get("project_id")
    ))

    # 统计资质证书数
    certs_data = _load_json(os.path.join(DATA_DIR, "certs.json"), {"certs": []})
    total_certs = len(certs_data.get("certs", []))

    # 统计即将到期的证书（90天内）
    expiring_soon = 0
    today = datetime.now()
    for cert in certs_data.get("certs", []):
        expire = cert.get("expire", "")
        status = cert.get("status", "")
        if status == "expired":
            continue
        if not expire:
            continue
        try:
            expire_date = datetime.strptime(expire, "%Y-%m-%d")
            diff = (expire_date - today).days
            if 0 <= diff <= 90:
                expiring_soon += 1
        except ValueError:
            pass

    return jsonify({
        "total_files": total_files,
        "analyzed_projects": analyzed_projects,
        "total_certs": total_certs,
        "expiring_soon": expiring_soon
    })


@projects_bp.route("/api/projects/fetch-url", methods=["POST"])
def fetch_announcement_url():
    """抓取公告网页内容（使用 Scrapling 作为 HTTP 爬虫引擎）"""
    body = request.get_json(silent=True) or {}
    url = body.get("url", "")

    if not url:
        return jsonify({"success": False, "message": "请提供网页链接"})

    try:
        from scrapling.fetchers import Fetcher
        import re

        # 1) 用 Scrapling 的 Fetcher 抓取页面（自动处理 UA、TLS 指纹、编码）
        page = Fetcher.get(url, stealthy_headers=True, timeout=20)

        # 2) 获取全文（Scrapling 已处理编码）
        all_text = str(page.get_all_text())

        # 3) CSS 选择器：优先从特定内容容器中查找（文本量最大的那块）
        candidate_selectors = [
            ".vF_detail_main",      # 中国政府采购网
            ".vT_detail_main",
            "div.content", "div.article-content", "div.detail-content",
            "div.main-content", "div.article-body", "div.detail-info",
            ".article-detail", ".detail-content", ".content-detail",
            ".zw", "article", "div#content", "div#main",
            ".article", ".detail", ".news", ".news-detail",
            ".notice-content", ".announcement-content",
            "table", ".container",
        ]

        best_text = ""
        best_label = "full_page_text"
        for sel in candidate_selectors:
            try:
                element = page.css(sel).get()
                if element is not None:
                    txt = element.get_text()
                    if isinstance(txt, str) and len(txt) > len(best_text) and len(txt) > 50:
                        best_text = txt
                        best_label = sel
            except Exception:
                continue

        if not best_text:
            best_text = all_text

        # 4) 清理多余空白行
        lines = [line.strip() for line in best_text.split("\n") if line and line.strip()]
        content_clean = "\n".join(lines)

        # 5) 在 content_clean 和 all_text 中用正则找 "项目名称"
        project_name_patterns = [
            r"项目名称[：:]\s*([^\n\r]+)",
            r"项目名称\s*[:：]?\s*([^\n\r]+)",
            r"项目名称[\s]*[:：]?[\s]*([^\n\r]+)",
            r"项目名称[：:]?[\s\n\r]+([^\n\r]{2,150})",
        ]

        title = ""
        matched_pattern = ""
        matched_raw_line = ""
        matched_source = ""

        for source_name, source_text in [("content_clean", content_clean), ("all_text", all_text)]:
            if title:
                break
            if not source_text:
                continue
            for i, pattern in enumerate(project_name_patterns):
                m = re.search(pattern, source_text)
                if m:
                    val = m.group(1).strip()
                    val = val.strip("《》<>''\"")
                    if len(val) > 100:
                        val = val[:100]
                    title = val
                    matched_pattern = f"Pattern {i+1}: {pattern}"
                    matched_raw_line = m.group(0)[:100]
                    matched_source = source_name
                    break

        # 6) 兜底：HTML <title> 标签
        fallback_title = ""
        if not title:
            try:
                html_title = page.css("title::text").get()
                if html_title:
                    fallback_title = str(html_title).strip()
                    title = fallback_title
                    matched_source = "html_title_fallback"
            except Exception:
                pass

        # 7) 调试信息：content 中含"项目名称"的行
        debug_lines = []
        for line in content_clean.split("\n"):
            if "项目名称" in line:
                debug_lines.append(line.strip())
            if len(debug_lines) >= 5:
                break

        debug_lines_all = []
        for line in all_text.split("\n"):
            if "项目名称" in line:
                debug_lines_all.append(line.strip())
            if len(debug_lines_all) >= 5:
                break

        return jsonify({
            "success": True,
            "title": title,
            "matched_pattern": matched_pattern,
            "matched_raw_line": matched_raw_line,
            "matched_source": matched_source,
            "fallback_from_html_title": fallback_title,
            "content_source": best_label,
            "content_clean_preview": content_clean[:500] if content_clean else "",
            "all_text_preview": all_text[:500] if all_text else "",
            "lines_with_project_name_in_content": debug_lines,
            "lines_with_project_name_in_all_text": debug_lines_all,
        })

    except Exception as e:
        return jsonify({"success": False, "message": f"抓取失败: {str(e)}"})


@projects_bp.route("/api/projects", methods=["POST"])
def create_project():
    """创建新项目（简化版：接收URL和文件，自动抓取信息）

    请求体（multipart/form-data）:
    - url: 公告网页地址（必填）
    - files: 上传的文件（可选，支持多文件）

    流程:
    1. 用 Scrapling 抓取 URL 内容，提取项目名称/编号/预算金额/招标人
    2. 创建项目记录
    3. 解析上传的文件（保存到 uploads/parsed 目录）
    4. 返回项目信息
    """
    url = request.form.get("url", "").strip()
    if not url:
        return jsonify({"success": False, "message": "请输入公告网页地址"}), 400

    # 1) 用 Scrapling 抓取 URL 内容
    scraped_info = {
        "project_name": "",
        "project_no": "",
        "budget_amount": "",
        "purchaser": "",
        "agency": "",
    }

    try:
        from scrapling.fetchers import Fetcher
        import re

        page = Fetcher.get(url, stealthy_headers=True, timeout=20)
        scraped_text = ""

        # 优先从内容容器提取
        candidate_selectors = [
            ".vF_detail_main", ".vT_detail_main",
            "div.content", "div.article-content", "div.detail-content",
            "div.main-content", "div.article-body", "div.detail-info",
            ".article-detail", ".detail-content", ".content-detail",
            ".zw", "article", "div#content", "div#main",
            ".article", ".detail", ".news", ".news-detail",
            ".notice-content", ".announcement-content",
            "table", ".container",
        ]
        for sel in candidate_selectors:
            try:
                el = page.css(sel).get()
                if el is not None:
                    txt = el.get_text()
                    if isinstance(txt, str) and len(txt) > 100:
                        scraped_text = txt
                        break
            except Exception:
                continue

        if not scraped_text:
            try:
                scraped_text = str(page.get_all_text())
            except Exception:
                scraped_text = ""

        # 用正则提取关键字段
        patterns = {
            "project_name": [r"项目名称[：:]\s*([^\n\r]+)", r"项目名称\s*[:：]?\s*([^\n\r]+)"],
            "project_no": [r"项目编号[：:]\s*([^\n\r]+)", r"招标编号[：:]\s*([^\n\r]+)"],
            "budget_amount": [r"预算金额[：:]\s*([^\n\r]+)", r"预算[：:]\s*([^\n\r]+)"],
            "purchaser": [r"采购人[：:]\s*([^\n\r]+)", r"招标人[：:]\s*([^\n\r]+)"],
            "agency": [r"代理机构[：:]\s*([^\n\r]+)", r"招标代理机构[：:]\s*([^\n\r]+)"],
        }

        for field, pats in patterns.items():
            for p in pats:
                m = re.search(p, scraped_text)
                if m and m.group(1).strip():
                    val = m.group(1).strip()
                    val = val.strip("《》<>''\"")
                    if len(val) > 150:
                        val = val[:150]
                    scraped_info[field] = val
                    break

        # 如果项目名称没找到，回退到 HTML title
        if not scraped_info["project_name"]:
            try:
                title_text = page.css("title::text").get()
                if title_text:
                    scraped_info["project_name"] = str(title_text).strip()[:150]
            except Exception:
                pass

    except Exception as e:
        # 抓取失败也允许创建，用 URL 作为名称兜底
        scraped_info["project_name"] = url[:60] + "..."

    # 2) 创建项目
    projects = load_projects()
    project_id = str(uuid.uuid4())[:8]
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    project = {
        "id": project_id,
        "name": scraped_info.get("project_name") or "未命名项目",
        "project_no": scraped_info.get("project_no", ""),
        "purchaser": scraped_info.get("purchaser", ""),
        "agency": scraped_info.get("agency", ""),
        "budget_amount": scraped_info.get("budget_amount", ""),
        "max_price": "",
        "bid_deadline": "",
        "opening_time": "",
        "announcement_url": url,
        "status": "pending",  # pending, parsing, extracting, completed
        "risk_count": 0,
        "high_risk_count": 0,
        "created_at": now,
        "updated_at": now,
    }

    projects.insert(0, project)
    save_projects(projects)

    # 初始化字段提取记录（空值，等待 AI 分析填充）
    _init_extracted_fields(project_id)

    # 3) 保存上传的文件
    project_upload_dir = os.path.join(UPLOADS_DIR, project_id)
    os.makedirs(project_upload_dir, exist_ok=True)

    files_records = []
    uploaded_files = request.files.getlist("file") + request.files.getlist("files")
    for file in uploaded_files:
        if not file or not file.filename:
            continue
        original_name = file.filename
        ext = os.path.splitext(original_name)[1].lower()
        file_id = str(uuid.uuid4())[:8]
        storage_name = f"{file_id}{ext}"
        storage_path = os.path.join(project_upload_dir, storage_name)
        file.save(storage_path)

        all_files = load_project_files()
        file_type = _get_file_type(ext)
        record = {
            "id": file_id,
            "project_id": project_id,
            "original_name": original_name,
            "storage_name": storage_name,
            "storage_path": storage_path,
            "file_type": file_type,
            "file_size": os.path.getsize(storage_path),
            "parse_status": "pending",
            "page_count": 0,
            "uploaded_at": now,
        }
        all_files.append(record)
        save_project_files(all_files)
        files_records.append(record)

    return jsonify({
        "success": True,
        "project": project,
        "files_count": len(files_records),
    })


@projects_bp.route("/api/projects/<project_id>", methods=["GET"])
def get_project(project_id):
    """获取项目详情"""
    projects = load_projects()
    project = next((p for p in projects if p["id"] == project_id), None)
    if not project:
        return jsonify({"success": False, "message": "项目不存在"}), 404

    # 获取关联文件
    files = load_project_files()
    project_files = [f for f in files if f.get("project_id") == project_id]

    # 获取提取字段
    fields = load_extracted_fields()
    valid_keys = set(FIELD_DEFINITIONS.keys())
    project_fields = [f for f in fields if f.get("project_id") == project_id and f.get("field_key") in valid_keys]

    # 兼容处理 1：如果该项目没有任何字段记录（可能是老项目），从 FIELD_DEFINITIONS 动态生成
    if not project_fields:
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for field_key, field_def in FIELD_DEFINITIONS.items():
            project_fields.append({
                "id": str(uuid.uuid4())[:8],
                "project_id": project_id,
                "field_key": field_key,
                "field_label": field_def["label"],
                "field_group": field_def["group"],
                "machine_value": "",
                "confirmed_value": "",
                "review_status": "pending",
                "confidence": 0.0,
                "source_file_id": "",
                "source_file_name": "",
                "created_at": now,
                "updated_at": now,
            })
    else:
        # 兼容处理 1b：补充新增的字段（如 announcement_url）
        existing_keys = {f["field_key"] for f in project_fields}
        for field_key, field_def in FIELD_DEFINITIONS.items():
            if field_key not in existing_keys:
                project_fields.append({
                    "id": str(uuid.uuid4())[:8],
                    "project_id": project_id,
                    "field_key": field_key,
                    "field_label": field_def["label"],
                    "field_group": field_def["group"],
                    "machine_value": "",
                    "confirmed_value": "",
                    "review_status": "pending",
                    "confidence": 0.0,
                    "source_file_id": "",
                    "source_file_name": "",
                    "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                })

    # 兼容处理 2：如果字段的 machine_value 为空，尝试从 project 对象获取 URL 抓取的值
    project_property_map = {
        "project_name": "name",
        "project_no": "project_no",
        "purchaser": "purchaser",
        "agency": "agency",
        "budget_amount": "budget_amount",
        "max_price": "max_price",
        "bid_deadline": "bid_deadline",
        "opening_time": "opening_time",
        "announcement_url": "announcement_url",
    }
    for f in project_fields:
        if not f.get("machine_value"):
            prop_key = project_property_map.get(f["field_key"])
            if prop_key and project.get(prop_key):
                f["machine_value"] = project[prop_key]

    # 获取风险项
    risks = load_risk_items()
    project_risks = [r for r in risks if r.get("project_id") == project_id]

    # 统计复核进度
    field_review_stats = {
        "total": len(project_fields),
        "pending": sum(1 for f in project_fields if f.get("review_status") == "pending"),
        "confirmed": sum(1 for f in project_fields if f.get("review_status") == "confirmed"),
        "modified": sum(1 for f in project_fields if f.get("review_status") == "modified"),
        "ignored": sum(1 for f in project_fields if f.get("review_status") == "ignored"),
    }

    risk_review_stats = {
        "total": len(project_risks),
        "pending": sum(1 for r in project_risks if r.get("review_status") == "pending"),
        "confirmed": sum(1 for r in project_risks if r.get("review_status") == "confirmed"),
        "ignored": sum(1 for r in project_risks if r.get("review_status") == "ignored"),
    }

    return jsonify({
        "project": project,
        "files": project_files,
        "fields": project_fields,
        "risks": project_risks,
        "field_review_stats": field_review_stats,
        "risk_review_stats": risk_review_stats,
    })


@projects_bp.route("/api/projects/<project_id>", methods=["PUT"])
def update_project(project_id):
    """更新项目信息"""
    body = request.get_json(silent=True) or {}
    projects = load_projects()

    for i, p in enumerate(projects):
        if p["id"] == project_id:
            # 更新允许修改的字段
            for key in ["name", "project_no", "purchaser", "agency", "budget_amount",
                        "max_price", "bid_deadline", "opening_time", "status", "owner", "note", "criteria_info"]:
                if key in body:
                    projects[i][key] = body[key]
            projects[i]["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            save_projects(projects)
            return jsonify({"success": True, "project": projects[i]})

    return jsonify({"success": False, "message": "项目不存在"}), 404


@projects_bp.route("/api/projects/<project_id>", methods=["DELETE"])
def delete_project(project_id):
    """删除项目及其关联数据"""
    errors = []

    try:
        # 删除项目记录
        projects = load_projects()
        projects = [p for p in projects if p["id"] != project_id]
        save_projects(projects)
    except Exception as e:
        errors.append(f"删除项目记录失败: {e}")

    try:
        # 删除关联文件记录
        files = load_project_files()
        files = [f for f in files if f.get("project_id") != project_id]
        save_project_files(files)
    except Exception as e:
        errors.append(f"删除文件记录失败: {e}")

    try:
        # 删除提取字段
        fields = load_extracted_fields()
        fields = [f for f in fields if f.get("project_id") != project_id]
        save_extracted_fields(fields)
    except Exception as e:
        errors.append(f"删除字段数据失败: {e}")

    try:
        # 删除风险项
        risks = load_risk_items()
        risks = [r for r in risks if r.get("project_id") != project_id]
        save_risk_items(risks)
    except Exception as e:
        errors.append(f"删除风险数据失败: {e}")

    # 删除上传的文件目录
    project_upload_dir = os.path.join(UPLOADS_DIR, project_id)
    if os.path.exists(project_upload_dir):
        try:
            shutil.rmtree(project_upload_dir)
        except Exception as e:
            errors.append(f"删除上传文件失败: {e}")

    # 删除解析结果目录
    project_parsed_dir = os.path.join(PARSED_DIR, project_id)
    if os.path.exists(project_parsed_dir):
        try:
            shutil.rmtree(project_parsed_dir)
        except Exception as e:
            errors.append(f"删除解析结果失败: {e}")

    # 删除关联的导出文件
    if os.path.exists(EXPORTS_DIR):
        for fname in list(os.listdir(EXPORTS_DIR)):
            if project_id in fname:
                fpath = os.path.join(EXPORTS_DIR, fname)
                try:
                    if os.path.isfile(fpath):
                        os.remove(fpath)
                except OSError as e:
                    errors.append(f"删除导出文件 {fname} 失败: {e}")

    if errors:
        return jsonify({"success": False, "message": "; ".join(errors)})

    return jsonify({"success": True})


def _init_extracted_fields(project_id):
    """初始化项目的字段提取记录"""
    fields = load_extracted_fields()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for field_key, field_def in FIELD_DEFINITIONS.items():
        fields.append({
            "id": str(uuid.uuid4())[:8],
            "project_id": project_id,
            "field_key": field_key,
            "field_label": field_def["label"],
            "field_group": field_def["group"],
            "machine_value": "",
            "confirmed_value": "",
            "source_file_id": "",
            "source_page_no": "",
            "source_excerpt": "",
            "confidence": 0,
            "review_status": "pending",
            "note": "",
            "created_at": now,
            "updated_at": now,
        })

    save_extracted_fields(fields)


# ── 文件上传与解析 ────────────────────────────────────────────
@projects_bp.route("/api/projects/<project_id>/files", methods=["POST"])
def upload_project_file(project_id):
    """上传项目文件"""
    if "file" not in request.files:
        return jsonify({"success": False, "message": "未选择文件"}), 400

    file = request.files["file"]
    if not file.filename:
        return jsonify({"success": False, "message": "文件名为空"}), 400

    # 创建项目上传目录
    project_upload_dir = os.path.join(UPLOADS_DIR, project_id)
    os.makedirs(project_upload_dir, exist_ok=True)

    # 生成唯一文件名
    original_name = file.filename
    ext = os.path.splitext(original_name)[1].lower()
    file_id = str(uuid.uuid4())[:8]
    storage_name = f"{file_id}{ext}"
    storage_path = os.path.join(project_upload_dir, storage_name)

    file.save(storage_path)

    # 判断文件类型
    file_type = _get_file_type(ext)

    # 保存文件记录
    files = load_project_files()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    file_record = {
        "id": file_id,
        "project_id": project_id,
        "original_name": original_name,
        "storage_name": storage_name,
        "storage_path": storage_path,
        "file_type": file_type,
        "file_size": os.path.getsize(storage_path),
        "parse_status": "pending",  # pending, parsing, done, failed
        "page_count": 0,
        "uploaded_at": now,
    }
    files.append(file_record)
    save_project_files(files)

    return jsonify({"success": True, "file": file_record})


@projects_bp.route("/api/projects/<project_id>/files", methods=["GET"])
def list_project_files(project_id):
    """获取项目文件列表"""
    files = load_project_files()
    project_files = [f for f in files if f.get("project_id") == project_id]
    return jsonify({"files": project_files})


@projects_bp.route("/api/projects/<project_id>/files/<file_id>", methods=["DELETE"])
def delete_project_file(project_id, file_id):
    """删除项目文件"""
    files = load_project_files()
    file_record = next((f for f in files if f["id"] == file_id and f["project_id"] == project_id), None)

    if file_record:
        # 删除物理文件
        if os.path.exists(file_record.get("storage_path", "")):
            os.remove(file_record["storage_path"])

        # 删除记录
        files = [f for f in files if f["id"] != file_id]
        save_project_files(files)

    return jsonify({"success": True})


@projects_bp.route("/api/projects/<project_id>/files/<file_id>/download", methods=["GET"])
def download_project_file(project_id, file_id):
    """下载项目文件"""
    files = load_project_files()
    file_record = next((f for f in files if f["id"] == file_id and f["project_id"] == project_id), None)

    if not file_record:
        return jsonify({"success": False, "message": "文件不存在"}), 404

    return send_file(file_record["storage_path"],
                     as_attachment=True,
                     download_name=file_record["original_name"])


def _get_file_type(ext):
    """根据扩展名判断文件类型"""
    type_map = {
        ".pdf": "PDF",
        ".doc": "DOC",
        ".docx": "DOCX",
        ".xls": "XLS",
        ".xlsx": "XLSX",
        ".zip": "ZIP",
        ".rar": "RAR",
        ".png": "IMAGE",
        ".jpg": "IMAGE",
        ".jpeg": "IMAGE",
        ".bmp": "IMAGE",
        ".webp": "IMAGE",
        ".txt": "TEXT",
        ".html": "HTML",
    }
    return type_map.get(ext, "OTHER")


# ── 字段提取与复核 ────────────────────────────────────────────
@projects_bp.route("/api/projects/<project_id>/fields", methods=["GET"])
def get_project_fields(project_id):
    """获取项目的字段提取结果"""
    fields = load_extracted_fields()
    valid_keys = set(FIELD_DEFINITIONS.keys())
    project_fields = [
        f for f in fields
        if f.get("project_id") == project_id and f.get("field_key") in valid_keys
    ]

    # 按分组组织
    grouped = {}
    for f in project_fields:
        group = f.get("field_group", "其他")
        if group not in grouped:
            grouped[group] = []
        grouped[group].append(f)

    return jsonify({"fields": project_fields, "grouped": grouped})


@projects_bp.route("/api/projects/<project_id>/fields/<field_id>", methods=["PUT"])
def update_project_field(project_id, field_id):
    """更新字段（人工复核）"""
    body = request.get_json(silent=True) or {}
    fields = load_extracted_fields()

    for i, f in enumerate(fields):
        if f["id"] == field_id and f["project_id"] == project_id:
            # 更新字段
            if "confirmed_value" in body:
                fields[i]["confirmed_value"] = body["confirmed_value"]
            if "review_status" in body:
                fields[i]["review_status"] = body["review_status"]
            if "note" in body:
                fields[i]["note"] = body["note"]
            fields[i]["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            save_extracted_fields(fields)
            return jsonify({"success": True, "field": fields[i]})

    return jsonify({"success": False, "message": "字段不存在"}), 404


# ── 风险识别与管理 ────────────────────────────────────────────
@projects_bp.route("/api/projects/<project_id>/risks", methods=["GET"])
def get_project_risks(project_id):
    """获取项目的风险清单"""
    risks = load_risk_items()
    project_risks = [r for r in risks if r.get("project_id") == project_id]

    # 筛选
    category = request.args.get("category", "")
    severity = request.args.get("severity", "")
    status = request.args.get("status", "")

    if category:
        project_risks = [r for r in project_risks if r.get("category") == category]
    if severity:
        project_risks = [r for r in project_risks if r.get("severity") == severity]
    if status:
        project_risks = [r for r in project_risks if r.get("review_status") == status]

    return jsonify({"risks": project_risks})


@projects_bp.route("/api/projects/<project_id>/risks/<risk_id>", methods=["PUT"])
def update_project_risk(project_id, risk_id):
    """更新风险项（人工复核）"""
    body = request.get_json(silent=True) or {}
    risks = load_risk_items()

    for i, r in enumerate(risks):
        if r["id"] == risk_id and r["project_id"] == project_id:
            if "review_status" in body:
                risks[i]["review_status"] = body["review_status"]
            if "note" in body:
                risks[i]["note"] = body["note"]
            if "severity" in body:
                risks[i]["severity"] = body["severity"]
            risks[i]["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            save_risk_items(risks)
            return jsonify({"success": True, "risk": risks[i]})

    return jsonify({"success": False, "message": "风险项不存在"}), 404


# ── 字段定义 API ────────────────────────────────────────────
@projects_bp.route("/api/field-definitions", methods=["GET"])
def get_field_definitions():
    """获取字段定义"""
    return jsonify({"definitions": FIELD_DEFINITIONS})


@projects_bp.route("/api/risk-categories", methods=["GET"])
def get_risk_categories():
    """获取风险类别"""
    return jsonify({"categories": RISK_CATEGORIES})


# ── 导出功能 ────────────────────────────────────────────────────
@projects_bp.route("/api/projects/<project_id>/export/fields", methods=["GET"])
def export_project_fields(project_id):
    """导出项目字段为 Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        return jsonify({"success": False, "message": "请先安装 openpyxl: pip install openpyxl"}), 500

    projects = load_projects()
    project = next((p for p in projects if p["id"] == project_id), None)
    if not project:
        return jsonify({"success": False, "message": "项目不存在"}), 404

    fields = load_extracted_fields()
    project_fields = [f for f in fields if f.get("project_id") == project_id]

    # 创建 Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "项目登记本"

    # 表头样式
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 写入标题
    ws.merge_cells('A1:H1')
    ws['A1'] = f"项目登记本 - {project.get('name', '')}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    # 写入表头
    headers = ["字段名称", "机器提取值", "人工确认值", "来源文件", "页码", "原文片段", "置信度", "复核状态"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    # 写入数据
    status_map = {"pending": "待复核", "confirmed": "已确认", "modified": "已修改", "ignored": "已忽略"}
    for row, field in enumerate(project_fields, 4):
        ws.cell(row=row, column=1, value=field.get("field_label", "")).border = thin_border
        ws.cell(row=row, column=2, value=field.get("machine_value", "")).border = thin_border
        ws.cell(row=row, column=3, value=field.get("confirmed_value", "")).border = thin_border
        ws.cell(row=row, column=4, value=field.get("source_file_name", "")).border = thin_border
        ws.cell(row=row, column=5, value=field.get("source_page_no", "")).border = thin_border
        ws.cell(row=row, column=6, value=field.get("source_excerpt", "")).border = thin_border
        ws.cell(row=row, column=7, value=field.get("confidence", 0)).border = thin_border
        ws.cell(row=row, column=8, value=status_map.get(field.get("review_status", ""), "")).border = thin_border

    # 调整列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 10

    # 保存文件
    export_filename = f"项目登记本_{project_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    export_path = os.path.join(EXPORTS_DIR, export_filename)
    wb.save(export_path)

    return send_file(export_path, as_attachment=True, download_name=export_filename)


@projects_bp.route("/api/projects/<project_id>/export/risks", methods=["GET"])
def export_project_risks(project_id):
    """导出风险清单为 Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        return jsonify({"success": False, "message": "请先安装 openpyxl: pip install openpyxl"}), 500

    projects = load_projects()
    project = next((p for p in projects if p["id"] == project_id), None)
    if not project:
        return jsonify({"success": False, "message": "项目不存在"}), 404

    risks = load_risk_items()
    project_risks = [r for r in risks if r.get("project_id") == project_id]

    # 创建 Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "废标风险清单"

    # 样式
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FDE2E2", end_color="FDE2E2", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 写入标题
    ws.merge_cells('A1:I1')
    ws['A1'] = f"废标风险清单 - {project.get('name', '')}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    # 写入表头
    headers = ["风险标题", "风险类别", "严重程度", "原文内容", "来源文件", "页码", "触发原因", "复核状态", "备注"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    # 写入数据
    severity_map = {"high": "高", "medium": "中", "low": "低", "unknown": "待确认"}
    status_map = {"pending": "待复核", "confirmed": "已确认", "ignored": "已忽略"}
    for row, risk in enumerate(project_risks, 4):
        ws.cell(row=row, column=1, value=risk.get("title", "")).border = thin_border
        ws.cell(row=row, column=2, value=risk.get("category", "")).border = thin_border
        ws.cell(row=row, column=3, value=severity_map.get(risk.get("severity", ""), "")).border = thin_border
        ws.cell(row=row, column=4, value=risk.get("original_text", "")).border = thin_border
        ws.cell(row=row, column=5, value=risk.get("source_file_name", "")).border = thin_border
        ws.cell(row=row, column=6, value=risk.get("source_page_no", "")).border = thin_border
        ws.cell(row=row, column=7, value=risk.get("trigger_reason", "")).border = thin_border
        ws.cell(row=row, column=8, value=status_map.get(risk.get("review_status", ""), "")).border = thin_border
        ws.cell(row=row, column=9, value=risk.get("note", "")).border = thin_border

    # 调整列宽
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 20

    # 保存文件
    export_filename = f"废标风险清单_{project_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    export_path = os.path.join(EXPORTS_DIR, export_filename)
    wb.save(export_path)

    return send_file(export_path, as_attachment=True, download_name=export_filename)


# ── 文件解析 API ────────────────────────────────────────────
@projects_bp.route("/api/projects/<project_id>/files/<file_id>/parse", methods=["POST"])
def parse_project_file(project_id, file_id):
    """解析项目文件"""
    files = load_project_files()
    file_record = next((f for f in files if f["id"] == file_id and f["project_id"] == project_id), None)

    if not file_record:
        return jsonify({"success": False, "message": "文件不存在"}), 404

    # 导入解析模块
    import sys
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    from backend.doc_parser import parse_file

    # 解析文件
    result = parse_file(file_record["storage_path"], file_record["file_type"])

    if result.get("error"):
        # 更新状态为失败
        for i, f in enumerate(files):
            if f["id"] == file_id:
                files[i]["parse_status"] = "failed"
                files[i]["parse_error"] = result["error"]
                break
        save_project_files(files)
        return jsonify({"success": False, "message": result["error"]})

    # 更新文件记录
    for i, f in enumerate(files):
        if f["id"] == file_id:
            files[i]["parse_status"] = "done"
            files[i]["page_count"] = result.get("page_count", 0)
            files[i]["parsed_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            break
    save_project_files(files)

    # 保存解析结果
    parsed_dir = os.path.join(PARSED_DIR, project_id)
    os.makedirs(parsed_dir, exist_ok=True)
    parsed_file = os.path.join(parsed_dir, f"{file_id}.json")
    with open(parsed_file, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    return jsonify({
        "success": True,
        "file": files[next(i for i, f in enumerate(files) if f["id"] == file_id)],
        "parse_result": {
            "page_count": result.get("page_count", 0),
            "text_length": len(result.get("text", ""))
        }
    })


@projects_bp.route("/api/projects/<project_id>/files/<file_id>/parsed", methods=["GET"])
def get_parsed_content(project_id, file_id):
    """获取解析后的文件内容"""
    parsed_file = os.path.join(PARSED_DIR, project_id, f"{file_id}.json")

    if not os.path.exists(parsed_file):
        return jsonify({"success": False, "message": "文件尚未解析"}), 404

    with open(parsed_file, "r", encoding="utf-8") as f:
        result = json.load(f)

    return jsonify({"success": True, "result": result})


# ── AI 提取 API ────────────────────────────────────────────
@projects_bp.route("/api/projects/<project_id>/extract", methods=["POST"])
def extract_project_info(project_id):
    """使用 AI 提取项目信息和识别风险"""
    # 检查项目是否存在
    projects = load_projects()
    project = next((p for p in projects if p["id"] == project_id), None)
    if not project:
        return jsonify({"success": False, "message": "项目不存在"}), 404

    # 获取项目文件
    files = load_project_files()
    project_files = [f for f in files if f.get("project_id") == project_id and f.get("parse_status") == "done"]

    if not project_files:
        return jsonify({"success": False, "message": "没有已解析的文件，请先上传并解析文件"}), 400

    # 导入 AI 提取模块
    import sys
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    from backend.ai_extractor import analyze_document_full

    # 合并所有文件的文本
    all_text = []
    for f in project_files:
        parsed_file = os.path.join(PARSED_DIR, project_id, f"{f['id']}.json")
        if os.path.exists(parsed_file):
            with open(parsed_file, "r", encoding="utf-8") as pf:
                parsed_data = json.load(pf)
                if parsed_data.get("text"):
                    all_text.append(f"【文件：{f['original_name']}】\n{parsed_data['text']}")

    if not all_text:
        return jsonify({"success": False, "message": "没有可用的文本内容"}), 400

    combined_text = "\n\n".join(all_text)

    # 更新项目状态为解析中
    for i, p in enumerate(projects):
        if p["id"] == project_id:
            projects[i]["status"] = "extracting"
            break
    save_projects(projects)

    # 执行 AI 分析
    result = analyze_document_full(combined_text, project_id)

    # 更新字段提取结果
    if result["field_success"]:
        _update_extracted_fields(project_id, result["fields"], project_files[0] if project_files else None)

    # 保存风险项
    if result["risk_success"]:
        _save_risk_items(project_id, result["risks"], project_files[0] if project_files else None)

    # 更新项目状态
    for i, p in enumerate(projects):
        if p["id"] == project_id:
            projects[i]["status"] = "parsed"
            projects[i]["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            # 更新风险统计
            risks = load_risk_items()
            project_risks = [r for r in risks if r.get("project_id") == project_id]
            projects[i]["risk_count"] = len(project_risks)
            projects[i]["high_risk_count"] = sum(1 for r in project_risks if r.get("severity") == "high")
            break
    save_projects(projects)

    return jsonify({
        "success": True,
        "field_success": result["field_success"],
        "risk_success": result["risk_success"],
        "errors": result.get("errors", []),
        "stats": {
            "fields_extracted": len([v for v in result.get("fields", {}).values() if v]),
            "risks_found": len(result.get("risks", []))
        }
    })


def _update_extracted_fields(project_id, extracted_values, source_file=None):
    """更新或创建字段提取结果
    - 对于已存在的字段：更新 machine_value（如果 AI 提取到了新值）
    - 对于不存在的字段：从 FIELD_DEFINITIONS 中创建新条目，填入 AI 提取值（即使为空也创建占位符）
    """
    fields = load_extracted_fields()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    source_file_id = source_file.get("id", "") if source_file else ""
    source_file_name = source_file.get("original_name", "") if source_file else ""

    # 构建已存在字段的索引：key = field_key
    existing_index = {}
    for i, f in enumerate(fields):
        if f["project_id"] == project_id:
            existing_index[f["field_key"]] = i

    for field_key, field_def in FIELD_DEFINITIONS.items():
        new_value = extracted_values.get(field_key, "")
        if field_key in existing_index:
            # 已存在的字段：有值才更新
            if new_value:
                idx = existing_index[field_key]
                fields[idx]["machine_value"] = new_value
                fields[idx]["confidence"] = 0.8
                fields[idx]["source_file_id"] = source_file_id
                fields[idx]["source_file_name"] = source_file_name
                fields[idx]["updated_at"] = now
        else:
            # 不存在的字段：创建新条目
            new_field = {
                "id": str(uuid.uuid4())[:8],
                "project_id": project_id,
                "field_key": field_key,
                "field_label": field_def["label"],
                "field_group": field_def["group"],
                "machine_value": new_value or "",
                "confirmed_value": "",
                "review_status": "pending",
                "confidence": 0.8 if new_value else 0.0,
                "source_file_id": source_file_id,
                "source_file_name": source_file_name,
                "created_at": now,
                "updated_at": now,
            }
            fields.append(new_field)

    save_extracted_fields(fields)


def _save_risk_items(project_id, risks, source_file=None):
    """保存风险项"""
    existing_risks = load_risk_items()
    # 删除该项目的旧风险项
    existing_risks = [r for r in existing_risks if r.get("project_id") != project_id]

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for risk in risks:
        risk_item = {
            "id": str(uuid.uuid4())[:8],
            "project_id": project_id,
            "title": risk.get("title", ""),
            "category": risk.get("category", "其他需关注条款"),
            "severity": risk.get("severity", "unknown"),
            "original_text": risk.get("original_text", ""),
            "trigger_reason": risk.get("trigger_reason", ""),
            "source_file_id": source_file.get("id", "") if source_file else "",
            "source_file_name": source_file.get("original_name", "") if source_file else "",
            "source_page_no": "",
            "section_title": "",
            "review_status": "pending",
            "note": "",
            "created_at": now,
            "updated_at": now,
        }
        existing_risks.append(risk_item)

    save_risk_items(existing_risks)


# ── 批量操作 API ────────────────────────────────────────────
@projects_bp.route("/api/projects/<project_id>/parse-all", methods=["POST"])
def parse_all_project_files(project_id):
    """解析项目的所有文件"""
    files = load_project_files()
    project_files = [f for f in files if f.get("project_id") == project_id and f.get("parse_status") != "done"]

    if not project_files:
        return jsonify({"success": True, "message": "没有需要解析的文件", "parsed_count": 0})

    import sys
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    from backend.doc_parser import parse_file

    parsed_count = 0
    errors = []

    for f in project_files:
        result = parse_file(f["storage_path"], f["file_type"])

        for i, file_rec in enumerate(files):
            if file_rec["id"] == f["id"]:
                if result.get("error"):
                    files[i]["parse_status"] = "failed"
                    files[i]["parse_error"] = result["error"]
                    errors.append(f"{f['original_name']}: {result['error']}")
                else:
                    files[i]["parse_status"] = "done"
                    files[i]["page_count"] = result.get("page_count", 0)
                    files[i]["parsed_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    parsed_count += 1

                    # 保存解析结果
                    parsed_dir = os.path.join(PARSED_DIR, project_id)
                    os.makedirs(parsed_dir, exist_ok=True)
                    parsed_file = os.path.join(parsed_dir, f"{f['id']}.json")
                    with open(parsed_file, "w", encoding="utf-8") as pf:
                        json.dump(result, pf, ensure_ascii=False, indent=2)
                break

    save_project_files(files)

    return jsonify({
        "success": True,
        "parsed_count": parsed_count,
        "total_count": len(project_files),
        "errors": errors
    })


@projects_bp.route("/api/projects/<project_id>/markdown", methods=["GET"])
def get_project_markdown(project_id):
    """获取项目的公告网页 MD 内容"""
    projects = load_projects()
    project = next((p for p in projects if p["id"] == project_id), None)
    if not project:
        return jsonify({"success": False, "message": "项目不存在"}), 404

    md_content = get_saved_markdown(project_id)
    if md_content:
        return jsonify({
            "success": True,
            "has_markdown": True,
            "content": md_content,
            "content_length": len(md_content),
            "source_url": project.get("announcement_url", "")
        })
    else:
        return jsonify({
            "success": True,
            "has_markdown": False,
            "content": "",
            "source_url": project.get("announcement_url", ""),
            "message": "尚未转换公告网页为MD，请在分析时自动转换或调用转换接口"
        })


@projects_bp.route("/api/projects/<project_id>/markdown/convert", methods=["POST"])
def convert_project_markdown(project_id):
    """手动触发将公告网页转换为 Markdown（可重复执行以刷新）"""
    projects = load_projects()
    project = next((p for p in projects if p["id"] == project_id), None)
    if not project:
        return jsonify({"success": False, "message": "项目不存在"}), 404

    announcement_url = project.get("announcement_url", "")
    if not announcement_url:
        return jsonify({"success": False, "message": "该项目没有配置公告网页链接"}), 400

    md_result = convert_url_to_markdown(announcement_url, project_id)

    if md_result["success"]:
        return jsonify({
            "success": True,
            "content_length": md_result["content_length"],
            "md_file_path": md_result.get("md_file_path", ""),
            "message": f"成功转换为Markdown，共 {md_result['content_length']} 字符"
        })
    else:
        return jsonify({
            "success": False,
            "error": md_result.get("error", "未知错误")
        }), 500


# ── 开始分析：解析 + AI提取 一条龙（含markitdown网页MD） ──────
@projects_bp.route("/api/projects/<project_id>/analyze", methods=["POST"])
def analyze_project(project_id):
    """一站式分析：markitdown转MD + 解析文件 + AI提取字段和风险

    新流程：
    1. 用 markitdown 将公告网页 URL 转换为 Markdown
    2. 解析所有上传的投标文件
    3. 将 MD内容 + 投标文件内容 合并同时发给大模型
    4. 大模型从多信息源中提取字段和识别风险（实现交叉匹配）
    """
    projects = load_projects()
    project = next((p for p in projects if p["id"] == project_id), None)
    if not project:
        return jsonify({"success": False, "message": "项目不存在"}), 404

    # ═══ Step 1: 使用 markitdown 将公告网页转换为 Markdown ═══
    md_content = ""
    md_result = {"success": False, "error": None}
    announcement_url = project.get("announcement_url", "")

    if announcement_url:
        md_result = convert_url_to_markdown(announcement_url, project_id)
        if md_result["success"]:
            md_content = md_result["markdown_content"]
            print(f"[markitdown] 成功转换公告网页为MD，长度: {md_result['content_length']} 字符")
        else:
            print(f"[markitdown] 转换失败: {md_result.get('error', '未知错误')}")
            # MD转换失败不阻断流程，继续用文件分析

    # ═══ Step 2: 解析所有未解析的投标文件 ═══
    files = load_project_files()
    project_files_all = [f for f in files if f.get("project_id") == project_id]
    files_to_parse = [f for f in project_files_all if f.get("parse_status") != "done"]

    parsed_count = 0
    parse_errors = []
    if files_to_parse:
        import sys as _sys_parse
        _sys_parse.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        from backend.doc_parser import parse_file

        for f in files_to_parse:
            result = parse_file(f["storage_path"], f["file_type"])
            for i, file_rec in enumerate(files):
                if file_rec["id"] == f["id"]:
                    if result.get("error"):
                        files[i]["parse_status"] = "failed"
                        files[i]["parse_error"] = result["error"]
                        parse_errors.append(f"{f['original_name']}: {result['error']}")
                    else:
                        files[i]["parse_status"] = "done"
                        files[i]["page_count"] = result.get("page_count", 0)
                        files[i]["parsed_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        parsed_count += 1
                        # 保存解析结果
                        parsed_dir = os.path.join(PARSED_DIR, project_id)
                        os.makedirs(parsed_dir, exist_ok=True)
                        parsed_file = os.path.join(parsed_dir, f"{f['id']}.json")
                        with open(parsed_file, "w", encoding="utf-8") as pf:
                            json.dump(result, pf, ensure_ascii=False, indent=2)
                    break

        save_project_files(files)

    # ═══ Step 3: 准备发送给大模型的完整内容（MD + 文件） ═══
    parsed_files = [f for f in files if f.get("project_id") == project_id and f.get("parse_status") == "done"]

    if not parsed_files and not md_content:
        return jsonify({
            "success": False,
            "message": "没有可解析的文件且公告网页也无法转换，无法进行 AI 分析",
            "parse_result": {"parsed_count": parsed_count, "total_count": len(project_files_all), "errors": parse_errors},
            "md_result": {"success": md_result.get("success", False), "error": md_result.get("error")}
        }), 400

    # 更新项目状态为分析中
    for i, p in enumerate(projects):
        if p["id"] == project_id:
            projects[i]["status"] = "analyzing"
            break
    save_projects(projects)

    # 构建多源内容：MD内容 + 投标文件内容
    all_text_parts = []

    # 3a. 添加 markitdown 转换的公告 MD 内容（优先级高，放在最前面）
    if md_content:
        all_text_parts.append(f"【公告网页内容（来源：{announcement_url}）】\n{md_content}")

    # 3b. 添加投标文件的解析内容
    for f in parsed_files:
        parsed_file = os.path.join(PARSED_DIR, project_id, f"{f['id']}.json")
        if os.path.exists(parsed_file):
            with open(parsed_file, "r", encoding="utf-8") as pf:
                parsed_data = json.load(pf)
                content_parts = []
                if parsed_data.get("text"):
                    content_parts.append(parsed_data["text"])
                # 同时包含表格内容（预算/金额等关键信息常出现在表格中）
                if parsed_data.get("tables"):
                    table_text_parts = []
                    for idx, table in enumerate(parsed_data["tables"]):
                        rows_text = []
                        for row in table:
                            if isinstance(row, list):
                                rows_text.append(" | ".join(str(cell) for cell in row))
                            else:
                                rows_text.append(str(row))
                        if rows_text:
                            table_text_parts.append("[表格" + str(idx + 1) + "]\n" + "\n".join(rows_text))
                    if table_text_parts:
                        content_parts.append("\n\n" + "\n\n".join(table_text_parts))
                combined_content = "\n".join(content_parts)
                all_text_parts.append(f"【投标文件：{f['original_name']}】\n{combined_content}")

    combined_text = "\n\n---\n\n".join(all_text_parts)

    # ═══ Step 4: AI 提取字段和风险（基于多源内容） ═══
    import sys as _sys_ai
    _sys_path_backup = _sys_ai.path[:]
    _sys_ai.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    from backend.ai_extractor import analyze_document_full

    ai_result = analyze_document_full(combined_text, project_id)
    _sys_ai.path = _sys_path_backup

    # 保存字段结果
    if ai_result["field_success"]:
        _update_extracted_fields(project_id, ai_result["fields"], parsed_files[0] if parsed_files else None)

    # 保存风险项
    if ai_result["risk_success"]:
        _save_risk_items(project_id, ai_result["risks"], parsed_files[0] if parsed_files else None)

    # ═══ Step 5: 提取评分标准（用于模拟评分） ═══
    criteria_info = None
    try:
        import sys as _sys_criteria
        _sys_criteria.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        from backend.routes.bid_score import _build_criteria_extraction_messages, _normalize_criteria, _parse_json_response
        from backend.llm_client import call_llm, get_all_enabled_model_configs
        from backend.ai_extractor import _call_llm_with_fallback

        messages = _build_criteria_extraction_messages("请提取招标文件中的评分标准", combined_text)
        target_model = None
        all_models = get_all_enabled_model_configs()
        for m in all_models:
            if m.get("id"):
                target_model = m
                break

        if target_model:
            response = call_llm(messages, temperature=0.1, max_tokens=4000, model_config=target_model)
        else:
            response, _ = _call_llm_with_fallback(messages, temperature=0.1, max_tokens=4000)

        parsed = _parse_json_response(response)
        criteria_info = _normalize_criteria(parsed)
    except Exception as e:
        print(f"[评分标准提取] 失败: {e}")

    # 更新项目状态为已完成
    for i, p in enumerate(projects):
        if p["id"] == project_id:
            projects[i]["status"] = "completed"
            projects[i]["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            risks = load_risk_items()
            project_risks = [r for r in risks if r.get("project_id") == project_id]
            projects[i]["risk_count"] = len(project_risks)
            projects[i]["high_risk_count"] = sum(1 for r in project_risks if r.get("severity") == "high")
            if criteria_info:
                projects[i]["criteria_info"] = criteria_info
            break
    save_projects(projects)

    return jsonify({
        "success": True,
        "md_result": {
            "success": md_result.get("success", False),
            "content_length": md_result.get("content_length", 0),
            "error": md_result.get("error")
        },
        "parse_result": {
            "parsed_count": parsed_count,
            "total_count": len(project_files_all),
            "errors": parse_errors
        },
        "ai_result": {
            "field_success": ai_result["field_success"],
            "risk_success": ai_result["risk_success"],
            "fields_extracted_count": len([v for v in ai_result.get("fields", {}).values() if v]),
            "risks_found_count": len(ai_result.get("risks", [])),
            "errors": ai_result.get("errors", [])
        }
    })
